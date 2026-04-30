import openpyxl, re
from collections import Counter
import os
P = os.path.expandvars(r'%TEMP%\Inc.xlsx')
wb=openpyxl.load_workbook(P, read_only=True, data_only=True)
s=wb['Incidenten']
codes=Counter(); samples={}
for r in s.iter_rows(values_only=True):
    body=str(r[3] or '')
    if 'intrekken' in body.lower(): continue
    for c in re.findall(r'\b\d{6}\b', body):
        codes[c]+=1
        if codes[c]<=2: samples.setdefault(c,[]).append(body[:120])

print("=== top codes ===")
for k,v in sorted(codes.items(), key=lambda x:-x[1])[:30]:
    print(f"{v:5d}  {k}  | {samples.get(k,[''])[0][:90]}")

print("\n=== suspected HA (3386/3364/3387/3344) lines ===")
for code in ['093386','093364','093387','093344']:
    print(f"--- {code} (count {codes.get(code,0)}) ---")
    n=0
    for r in s.iter_rows(values_only=True):
        body=str(r[3] or '')
        if code in body:
            print(' ', body[:150]); n+=1
            if n>=4: break

print("\n=== natuurbrand patterns with TS code ===")
n=0
for r in s.iter_rows(values_only=True):
    body=str(r[3] or '')
    if any(w in body.lower() for w in ['heide','bos','natuur','duin']) and re.search(r'\b09(?:34[14]|34[34]1|3441|3341)\b', body):
        print(' ', body[:170]); n+=1
        if n>=10: break
